import openpyxl
from openpyxl.styles import Font, Alignment

class ReporteExcel:
    def __init__(self, nombre_archivo="Auditoria_Nomina.xlsx"):
        self.nombre_archivo = nombre_archivo
        # Columnas extendidas para auditoría
        self.columnas = ["CENTRO", "GREMIO", "TRABAJADOR", "EMPRESA", "CONCEPTO", "NRO_RECIBO", "TRAB_REF_PDF", "EMPR_REF_PDF"]
        self.wb = openpyxl.Workbook()
        self.ws_base = self.wb.active
        self.ws_base.title = "BASE"
        self.ws_cuadre = self.wb.create_sheet("Cuadre y Verificación")

    def generar_reporte(self, datos):
        if not datos: return False
        self.ws_base.append(self.columnas)
        for fila in datos:
            row = [
                fila["CENTRO"], fila["GREMIO"], fila["TRABAJADOR"], fila["EMPRESA"], 
                fila["CONCEPTO"], fila["NRO_RECIBO"], fila["TRAB_ORIGINAL"], fila["EMPR_ORIGINAL"]
            ]
            self.ws_base.append(row)
            # Formato contable para Excel
            idx = self.ws_base.max_row
            self.ws_base[f"C{idx}"].number_format = '#,##0.00'
            self.ws_base[f"D{idx}"].number_format = '#,##0.00'
            
        self._configurar_cuadre(datos)
        self.wb.save(self.nombre_archivo)
        return True
    
    def _configurar_cuadre(self, datos):
        self.ws_cuadre.append(["CONCEPTO", "TIPO", "TOTAL EXCEL", "ESTADO"])
        conceptos = sorted(list(set(d["CONCEPTO"] for d in datos)))
        row_idx = 2
        for concepto in conceptos:
            for tipo in ["TRABAJADOR", "EMPRESA"]:
                col_suma = "C" if tipo == "TRABAJADOR" else "D"
                f_excel = f"=SUMIF('BASE'!E:E, A{row_idx}, 'BASE'!{col_suma}:{col_suma})"
                self.ws_cuadre.append([concepto, tipo, f_excel, "PENDIENTE"])
                self.ws_cuadre[f"C{row_idx}"].number_format = '#,##0.00'
                row_idx += 1