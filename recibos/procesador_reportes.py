import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

def generar_reportes_consolidados(ruta_excel, nombre_hoja_base):
    meses = ["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]
    mes_actual = meses[datetime.now().month - 1]
    anio_actual = datetime.now().year
    
    nombre_ret = f"RETENCIONES {mes_actual} {anio_actual}"
    nombre_casas = f"CASAS COMERCIALES {mes_actual} {anio_actual}"

    df = pd.read_excel(ruta_excel, sheet_name=nombre_hoja_base)
    df = df[pd.to_numeric(df.iloc[:, 3], errors='coerce').notna()]

    grupos_retenciones = [
        {"busqueda": "S.S.O.", "nombre_normal": "S.S.O. (4%) 210", "nombre_deuda": "S.S.O. (DEUDA) 210.1.1"},
        {"busqueda": "PERDIDA", "nombre_normal": "PERDIDA INVOLUNTARIA DEL EMPLEO (P.I.E.) 244", "nombre_deuda": "PERDIDA INVOLUNTARIA DEL EMPLEO (DEUDA) 244.2"},
        {"busqueda": "PENSION", "nombre_normal": "FONDO PENSIONES (JUBILACION) 301", "nombre_deuda": "FONDO PENSIONES JUBILACION (DEUDA) 301.2"},
        {"busqueda": "VIV", "nombre_normal": "FONDO DE AHORRO OBLIGATORIO PARA LA VIV 245", "nombre_deuda": "FONDO DE AHORRO OBLIGATORIO PARA LA VIV (DEUDA) 245.3"}
    ]

    casas_comerciales = [
        ("COLEGIO BIOANALISTA", "COLEGIO BIOANALISTA 255"),
        ("NUTRIC", "COLEGIO NUTRIC. Y DIET. VZLA 286"),
        ("COLEGIO ENFERMERAS", "COLEGIO ENFERMERAS 224"),
        ("CARABOBO", "COLEGIO DE ENFERMERAS ESTADO CARABOBO 967"),
        ("MIRANDA", "COLEGIO DE ENFERMERA EDO. MIRANDA 603"),
        ("DESC", "DESC. DIA(S) NO LABORADO(S) 179"),
        ("INPRENFERMERA", "DELEGACIÓN REGIONAL INPRENFERMERA ARAGUA 370"),
        ("CAHORMINSAS", "CAHORMINSAS 212"),
        ("FUNERARIOS CAHORMINSAS", "SERVICIOS FUNERARIOS CAHORMINSAS 269"),
        ("CAEMINSA", "CAJA DE AHORRO CAEMINSA 238"),
        ("PRESTAMO CAJA", "PRESTAMO CAJA DE AHORRO 223"),
        ("UNICO DE TRAB", "SINDICATO UNICO DE TRAB.DE LA SALUD Y S 233"),
        ("OSBESS", "SINDICATO OSBESS ARAGUA 978"),
        ("ANESTECIOLOGOS", "SOCIEDAD ANESTECIOLOGOS 257"),
        ("SINBOPROENF", "SINBOPROENF 321"),
        ("SUNEP-SAS", "SUNEP-SAS 213"),
        ("FENASISTRASALUD", "FENASISTRASALUD 513"),
        ("SISTRASALUD", "SISTRASALUD 583"),
        ("SAPTRASEZ", "SAPTRASEZ 964"),
        ("SITRASSS-MIRANDA", "SITRASSS-MIRANDA 581"),
        ("ASUNAJUPENSAPROSO", "ASUNAJUPENSAPROSO 799"),
        ("TRIBUNAL DE PROTECCION", "TRIBUNAL DE PROTECCION DE NIÑOS, NIÑAS 807"),
        ("TRIBUNAL (PERMANENTE)", "TRIBUNAL (PERMANENTE) 215")
    ]

    wb = load_workbook(ruta_excel)

    def escribir_hoja(nombre_hoja, lista_config, modo_retenciones=False):
        if nombre_hoja in wb.sheetnames: del wb[nombre_hoja]
        ws = wb.create_sheet(nombre_hoja)
        fila_actual = 1
        for item in lista_config:
            busqueda = item["busqueda"] if modo_retenciones else item[0]
            mask = df.iloc[:, 0].str.contains(busqueda, case=False, na=False)
            sub = df[mask]
            if not sub.empty:
                inicio = fila_actual
                for _, row in sub.iterrows():
                    if modo_retenciones:
                        nombre_mostrar = item["nombre_deuda"] if "DEUDA" in str(row.iloc[0]).upper() else item["nombre_normal"]
                    else:
                        nombre_mostrar = item[1]
                    ws.cell(row=fila_actual, column=1, value=nombre_mostrar)
                    ws.cell(row=fila_actual, column=2, value=row.iloc[1])
                    ws.cell(row=fila_actual, column=3, value=float(row.iloc[3]))
                    fila_actual += 1
                ws.cell(row=fila_actual, column=2, value="TOTAL:")
                ws.cell(row=fila_actual, column=3, value=f"=SUM(C{inicio}:C{fila_actual-1})")
                fila_actual += 4

    escribir_hoja(nombre_ret, grupos_retenciones, True)
    escribir_hoja(nombre_casas, casas_comerciales, False)
    wb.save(ruta_excel)