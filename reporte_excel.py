import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

class ReporteExcel:
    def __init__(self, nombre_archivo="Auditoria_Nomina.xlsx", columnas=None):
        self.nombre_archivo = nombre_archivo
        self.columnas = columnas or ["CENTRO", "GREMIO", "TRABAJADOR", "EMPRESA", "CONCEPTO", "NRO_RECIBO"]
        self.wb = openpyxl.Workbook()
        
        # Pestaña 1: Base de Datos
        self.ws_base = self.wb.active
        self.ws_base.title = "Base de Datos"
        
        # Pestaña 2: Cuadre y Verificación
        self.ws_cuadre = self.wb.create_sheet("Cuadre y Verificación")

    def generar_reporte(self, datos):
        if not datos: return False
        
        # Usamos las columnas definidas
        self.ws_base.append(self.columnas)
        for fila in datos:
            # Construimos la fila dinámicamente según el orden de las columnas
            row = [fila.get(col, "") for col in self.columnas]
            # (Opcional: lógica para convertir a float si el nombre de columna es TRABAJADOR/EMPRESA)
            self.ws_base.append(row)
            
        self._configurar_cuadre(datos)
        self.wb.save(self.nombre_archivo)
        return True
    
    def _configurar_cuadre(self, datos):
        # Encabezados según tu archivo
        headers_cuadre = ["CONCEPTO", "TIPO", "TOTAL PDF", "TOTAL EXCEL", "DIFERENCIA", "ESTADO"]
        self.ws_cuadre.append(headers_cuadre)
        
        # Estilo para encabezados
        for cell in self.ws_cuadre[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        conceptos = sorted(list(set(d["CONCEPTO"] for d in datos)))
        row_idx = 2

        for concepto in conceptos:
            for tipo in ["TRABAJADOR", "EMPRESA"]:
                # Determinamos la columna a sumar en 'Base de Datos' (C para Trabajador, D para Empresa)
                col_suma = "C" if tipo == "TRABAJADOR" else "D"
                
                # FÓRMULAS
                # TOTAL EXCEL: SUMIF(RangoConceptos, EsteConcepto, RangoMontos)
                f_excel = f"=SUMIF('Base de Datos'!E:E, A{row_idx}, 'Base de Datos'!{col_suma}:{col_suma})"
                
                # DIFERENCIA: TOTAL PDF - TOTAL EXCEL
                f_dif = f"=C{row_idx}-D{row_idx}"
                
                # ESTADO: Si la diferencia es 0 (o muy cercana), "CUADRA", sino "REVISAR"
                f_estado = f'=IF(ABS(E{row_idx})<0.01, "CUADRA", "REVISAR")'

                self.ws_cuadre.append([
                    concepto,   # A
                    tipo,       # B
                    0,          # C - TOTAL PDF (Para ser llenado manualmente)
                    f_excel,    # D - TOTAL EXCEL
                    f_dif,      # E - DIFERENCIA
                    f_estado    # F - ESTADO
                ])
                row_idx += 1

        # Ajuste de ancho de columnas básico
        self.ws_cuadre.column_dimensions['A'].width = 30
        self.ws_cuadre.column_dimensions['B'].width = 15
        self.ws_cuadre.column_dimensions['F'].width = 15